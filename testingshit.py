from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('teste.xlsx')
ws = wb.active

Row = 1
Storage = []

for col in range(1, 49):
    char = get_column_letter(col)
    Storage.append(ws[char + str(Row)].value)

for col in range(1,49):
    char = get_column_letter(col)
    ws[char + str(Row+1)] = Storage[col-1]

print(Storage)




