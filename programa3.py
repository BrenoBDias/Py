from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('teste.xlsx')
ws = wb.active

MC = 49
MR = 743
row=1
SKU = 'B'
Estoque = 'AM'
Price = 'J'

pronto =  False
while pronto == False:
    pronto = True

    for row in range(2, MR):

        if ws[SKU+str(row)].value == None: #se SKU é vazio print "none"
            print('none')
        elif ws[SKU+str(row)].value == ws[SKU+str(row + 1)].value: #se SKU = SKU de baixo
            pronto = False
            try:
                Storage = [] #cria/esvazia lista vazia para guardar a linha inteira
                PStore = ws[Price+str(row + 1)].value #guarda preço de baixo antes da substituição
                EStore = ws[Estoque+str(row + 1)].value #guarda o estoque de baixo antes da soma
                EStore1 = ws[Estoque+str(row)].value

                for col in range(1, 49): #guarda cada célula da linha 'row' como um item da lista 'Storage'
                    char = get_column_letter(col)
                    Storage.append(ws[char + str(row)].value)  

                for col in range(1,49):  #substitui cada célula da linha 'row' por um item da lista 'Storage'
                    char = get_column_letter(col)
                    ws[char + str(row+1)] = Storage[col-1]
                try:
                    if EStore == None:
                        EStore = 0
                    if EStore1 == None:
                        EStore1 = 0
                    ws[Estoque+str(row + 1)].value = EStore1 + EStore #soma o estoque de EStore ao estoque de baixo (depois da substituição)

                except:
                    print('erro de célula')

                if ws[Price+str(row)].value > PStore: #se preço guardado < preço
                    ws[Price+str(row + 1)] = PStore #substitui preço de baixo por preço
                ws.delete_rows(row)
                print('ação feita')
            except:
                print('erro')
                pronto = True
        else:
            print('ação não necessária')

wb.save('teste.xlsx')

