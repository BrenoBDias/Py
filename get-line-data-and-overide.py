from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('name.xlsx')
ws = wb.active

row = 1
row_storage = []

for col in range(1, 49):
    char = get_column_letter(col)
    row_storage.append(ws[char + str(row)].value)

for col in range(1,49):
    char = get_column_letter(col)
    ws[char + str(row+1)] = row_storage[col-1]

print(row_storage)




